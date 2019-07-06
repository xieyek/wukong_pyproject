# -*- encoding:utf-8 -*-
# 请求方式封装
import time
import openpyxl

from BASIC_FLOW.flow import Flow
from BASIC_FLOW.member_regulation import MemberRegulation
from BASIC_FLOW.relationship import Relationship
from Common import var_operation


class File_Helper():

    def write_position(self, length):
        p = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
             "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z",
             "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN",
             "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ"
             ]
        return p[length-1]

    def read_xlsx(self, workbook_path="../case_file/member_case_create_data.xlsx",
                  save_path=r'../Report/case%r.xlsx' % (time.strftime('%Y-%m-%d.%H-%M-%S', time.localtime()))):
        print("用例开始测试")
        nCount = 0  # 用例总数计数器
        nFail = 0  # 失败用例数 计数器
        # 获取工作簿文件对象
        workbook = openpyxl.load_workbook(workbook_path)
        # 获取文件的工作表
        sheetsname = workbook.sheetnames
        # 获取表对象
        for sheetname in sheetsname:
            sheet = workbook[sheetname]
            # 定义第一行
            r = 0
            # 循环每一行
            for row in sheet.rows:
                r += 1
                # 判断是否第一行，是就跳过本次循环
                if r == 1:
                    continue
                case_data = []
                for cell in row:
                    case_data.append(cell.value)
                print(case_data)
                try:
                    print("进入调用方法阶段")
                    if sheetname == "平级超越与任务升级":
                        print("平级超越与任务升级%d, %d, %d, %d" %(int(case_data[16]), int(case_data[17]), int(case_data[20]), int(case_data[21])))
                        MemberRegulation().member_promotion_service_fee(int(case_data[16]), int(case_data[17]), int(case_data[20]), int(case_data[21]))
                    elif sheetname == "补包规则":
                        print("补包规则%d, %d, %d, %d" %(int(case_data[16]), int(case_data[17]), int(case_data[18]), int(case_data[20])))
                        MemberRegulation().member_fill_package(int(case_data[16]), int(case_data[17]), int(case_data[18]), int(case_data[20]))
                    elif sheetname == "各层代理购包价格验证":
                        print("各层代理购包价格验证%d, %d" %(int(case_data[16]), int(case_data[20])))
                        MemberRegulation().member_crad_price(int(case_data[16]), int(case_data[20]))
                    elif sheetname == "销售利润返佣":
                        print("销售利润返佣%d, %d, %d" %(int(case_data[16]), int(case_data[17]), int(case_data[18])))
                        MemberRegulation().member_sales_commission(int(case_data[16]), int(case_data[17]), int(case_data[18]))
                    else:
                        raise NameError("找不到对应的方法")
                    write_position = self.write_position(len(case_data))
                    sheet[write_position + str(r)] = "pass"
                    nCount += 1
                    time.sleep(2)
                except Exception as a:
                    print(a)
                    write_position = self.write_position(len(case_data))
                    sheet[write_position + str(r)] = "报错内容：%r"%a
                    nFail += 1
                    nCount += 1
                    time.sleep(2)

        workbook.save(save_path)
        pass_rate = (nCount-nFail)/nCount
        if pass_rate !=1:
            raise NameError("用例通过率只有%r"%pass_rate)

    def read_xlsx_create_data(self, workbook_path="../case_file/member_case.xlsx",
                              save_path=r'../case_file/member_case_create_data.xlsx'):
        # product_id = Flow.product_audit_dailysale(var_operation.operation_token)
        product_id = 1455
        print("开始创建数据")
        nCount = 0  # 用例总数计数器
        nFail = 0  # 失败用例数 计数器
        # 获取工作簿文件对象
        workbook = openpyxl.load_workbook(workbook_path)
        # 获取文件的工作表
        sheetsname = workbook.sheetnames
        # 获取表对象
        for sheetname in sheetsname:
            sheet = workbook[sheetname]
            # 定义第一行
            r = 0
            # 循环每一行
            for row in sheet.rows:
                r += 1
                # 判断是否第一行，是就跳过本次循环
                if r == 1:
                    continue
                case_data = []
                for cell in row:
                    case_data.append(cell.value)
                print(case_data)
                try:
                    res = Relationship.create_relation(case_data[0], case_data[1], case_data[2], case_data[3], case_data[4],
                                                 case_data[5], case_data[6], case_data[7], case_data[8], case_data[9],
                                                 case_data[10], case_data[11], case_data[12], case_data[13], case_data[14], case_data[15], product_id)
                    write_position = self.write_position(len(case_data))
                    sheet[write_position + str(r)] = "pass"
                    sheet["Q" + str(r)] = res[0]
                    sheet["R" + str(r)] = res[1]
                    sheet["S" + str(r)] = res[2]
                    sheet["T" + str(r)] = res[3]
                    nCount += 1
                    time.sleep(2)
                except Exception as a:
                    print(a)
                    write_position = self.write_position(len(case_data))
                    sheet[write_position + str(r)] = "报错内容：%r"%a
                    nFail += 1
                    nCount += 1
                    time.sleep(2)

        workbook.save(save_path)
        pass_rate = (nCount-nFail)/nCount
        if pass_rate !=1:
            raise NameError("用例通过率只有%r"%pass_rate)


# f = File_Helper()
# f.read_xlsx()
# f.read_xlsx_create_data()



